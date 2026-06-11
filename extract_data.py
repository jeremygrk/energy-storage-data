import openpyxl, json
from datetime import datetime

wb = openpyxl.load_workbook('data.xlsx', data_only=True)
all_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    if sheet_name == '德国储能装机':
        # Two header rows: row 1 is merged title, row 2 has actual column names
        headers = []
        for cell in ws[2]:
            val = cell.value
            headers.append(str(val) if val is not None else '')
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            clean_row = []
            for i, c in enumerate(row):
                if c is None:
                    clean_row.append(None)
                elif isinstance(c, datetime):
                    # Convert "2022-01-01" → "22年1月"
                    clean_row.append(f"{str(c.year)[2:]}年{c.month}月")
                elif isinstance(c, (int, float)):
                    clean_row.append(round(c, 4))
                elif isinstance(c, str) and c.startswith('='):
                    clean_row.append(None)
                else:
                    clean_row.append(str(c))
            rows.append(clean_row)
        all_data[sheet_name] = {'headers': headers, 'rows': rows}
        print(f'{sheet_name}: {len(rows)} rows, {len(headers)} cols')

    elif sheet_name == '美国储能装机':
        # Wide format: years in rows, months in columns
        # Cols B-M (1-12): 功率 MW, Col N (13): 全年
        # Cols Q-AB (17-28): 容量 MWh, Col AC (29): 全年
        combined_rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            year_val = row[0]
            if year_val is None:
                continue
            year = int(year_val) if isinstance(year_val, (int, float)) else year_val
            year_short = str(year)[2:] if len(str(year)) == 4 else str(year)
            for m in range(12):
                mw = row[m + 1]
                mwh = row[m + 17]
                mw_val = round(float(mw), 4) if mw is not None else None
                mwh_val = round(float(mwh), 4) if mwh is not None else None
                combined_rows.append([f"{year_short}年{m+1}月", mw_val, mwh_val])
        all_data['美国储能装机'] = {
            'headers': ['', '装机功率(MW)', '装机容量(MWh)'],
            'rows': combined_rows
        }
        print(f'美国储能装机: {len(combined_rows)} rows')

    elif sheet_name == '意大利储能装机':
        # Standard format with quarterly labels like "2019Q1"
        headers = []
        for cell in ws[1]:
            val = cell.value
            headers.append(str(val) if val is not None else '')
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            clean_row = []
            for i, c in enumerate(row):
                if c is None:
                    clean_row.append(None)
                elif isinstance(c, str) and c.startswith('='):
                    clean_row.append(None)
                elif isinstance(c, (int, float)):
                    clean_row.append(round(c, 4))
                else:
                    val_str = str(c)
                    # Reformat "2019Q1" → "19年Q1"
                    if len(val_str) == 6 and val_str[4] == 'Q' and val_str[:4].isdigit():
                        val_str = val_str[2:4] + '年Q' + val_str[5]
                    clean_row.append(val_str)
            rows.append(clean_row)
        all_data[sheet_name] = {'headers': headers, 'rows': rows}
        print(f'{sheet_name}: {len(rows)} rows, {len(headers)} cols')

    elif sheet_name == '逆变器出口数据':
        # Date format "202001" (YYYYMM, stored as int), last row is 合计
        headers = []
        for cell in ws[1]:
            val = cell.value
            headers.append(str(val) if val is not None else '')
        rows = []
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None or str(c).strip() == '' for c in raw_row):
                continue
            # Check for 合计 row
            first_val = raw_row[0]
            if first_val is not None and str(first_val).strip() in ('合计', '合計', '總計', '总计'):
                continue
            clean_row = []
            for i, c in enumerate(raw_row):
                if c is None:
                    clean_row.append(None)
                elif isinstance(c, str) and c.startswith('='):
                    clean_row.append(None)
                elif i == 0 and isinstance(c, (int, float)):
                    # Convert date int 202001 → "20年1月"
                    s = str(int(c))
                    if len(s) == 6:
                        clean_row.append(s[2:4] + '年' + str(int(s[4:6])) + '月')
                    else:
                        clean_row.append(s)
                elif isinstance(c, (int, float)):
                    clean_row.append(round(c, 4))
                else:
                    clean_row.append(str(c))
            rows.append(clean_row)
        all_data[sheet_name] = {'headers': headers, 'rows': rows}
        print(f'{sheet_name}: {len(rows)} rows, {len(headers)} cols')

    else:
        # Standard format: row 1 = headers, rows 2+ = data
        headers = []
        for cell in ws[1]:
            val = cell.value
            headers.append(str(val) if val is not None else '')
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            clean_row = []
            for c in row:
                if c is None:
                    clean_row.append(None)
                elif isinstance(c, str) and c.startswith('='):
                    clean_row.append(None)
                elif isinstance(c, (int, float)):
                    clean_row.append(round(c, 4))
                else:
                    clean_row.append(str(c))
            rows.append(clean_row)
        all_data[sheet_name] = {'headers': headers, 'rows': rows}
        print(f'{sheet_name}: {len(rows)} rows, {len(headers)} cols')

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print('\nDone — data.json updated with all sheets.')
